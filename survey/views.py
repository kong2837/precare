import csv
import re
import urllib.parse
from tempfile import NamedTemporaryFile
from typing import Any

from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models.query import QuerySet
from django.http import HttpResponse
from django.shortcuts import render
from django.views import View
from django.views.generic.detail import DetailView
from django.views.generic.edit import ProcessFormView
from django.views.generic.list import ListView
from openpyxl import Workbook

import survey.utils as utils
from accounts.views import SuperuserRequiredMixin
from huami.models import HuamiAccount
from survey.models import Survey, Question, UserSurvey, Reply, SurveyQuestion, Answer
from django.core.exceptions import FieldError


# Create your views here.

class MyLoginRequiredMixin(LoginRequiredMixin):
    login_url = "/accounts/login/"
    redirect_field_name = "redirect_to"


class SurveyListView(MyLoginRequiredMixin, ListView):
    """참여 가능한 설문 목록을 제공하는 클래스 기반 뷰"""
    template_name = 'survey/survey_list.html'
    model = Survey

    def get_queryset(self):
        custom_order = [
            "[연구시작시]연구 참여자 기초 건강 설문",
            "[상시]QUIPP",
            "[상시]조기진통위험 10문항",
            "[상시]임신스트레스 10문항",
            "[연구종료시]만족도 조사",
        ]
        surveys = Survey.objects.exclude(
            title__in=["[상시]QUIPP 유증상", "[상시]QUIPP 무증상"]
        )
        surveys_sorted = sorted(
            surveys,
            key=lambda s: custom_order.index(s.title) if s.title in custom_order else 999
        )
        return surveys_sorted



class SurveyDetailView(MyLoginRequiredMixin, DetailView):
    """설문에 대한 구체적인 정보를 제공하는 클래스 기반 뷰
    TODO 구현예정
    """
    template_name = 'survey/survey_detail.html'
    model = Survey


class UserSurveyListAdminView(SuperuserRequiredMixin, ListView):
    """다른 유저가 작성한 설문 결과들을 제공하는 클래스 기반 뷰
    """
    template_name = 'survey/user_survey_list.html'
    paginate_by = 5
    model = UserSurvey

    def get_queryset(self):
        queryset = UserSurvey.objects.filter(user__pk=self.kwargs['user_pk'],
                                             survey__pk=self.kwargs['survey_pk']).order_by('create_at')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_pk'])
        return context


class SurveyListAdminView(SuperuserRequiredMixin, ListView):
    """다른 유저가 작성한 설문들에 대한 정보를 제공하는 클래스 기반 뷰
    """
    template_name = 'survey/user_survey_list_admin.html'
    paginate_by = 5
    model = Survey

    def get_queryset(self) -> QuerySet[Survey]:
        queryset = Survey.objects.filter(users__pk=self.kwargs.get('pk')).distinct()
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user'] = get_user_model().objects.get(pk=self.kwargs.get('pk'))
        return context


class UserSurveyCsvView(SuperuserRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user_surveys = UserSurvey.objects.filter(user__pk=self.kwargs.get('user_pk'),
                                                 survey__pk=self.kwargs.get('survey_pk'))

        filename = "-".join((HuamiAccount.objects.get(user_id=self.kwargs.get('user_pk')).fullname
                             , Survey.objects.get(pk=self.kwargs.get('survey_pk')).title))
        response = HttpResponse(headers={
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename="{filename}.csv"'})

        file = csv.writer(response)
        head_line = ['pk']

        for question in Survey.objects.get(pk=self.kwargs.get('survey_pk')).questions.all():
            head_line.append(question.title)
        file.writerow(head_line)

        for user_survey in user_surveys:
            line = []
            for reply in user_survey.replies.all():
                line.append(reply.content)
            file.writerow(line)

        return response


class UserSurveyListView(MyLoginRequiredMixin, ListView):
    """설문 작성 결과를 제공하는 클래스 기반 뷰
    """
    template_name = 'survey/user_survey_list.html'
    paginate_by = 5

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """survey 정보를 추가한 context_data 반환

        Returns:
            dict[str, Any]: context_data를 반환
        """
        context = super().get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(**self.kwargs)
        return context

    def get_queryset(self) -> QuerySet[UserSurvey]:
        """UserSurvey에 대한 object_list를 반환
        현재 로그인 된 User가 선택한 Survey에 대한 UserSurvey들을 반환
        Returns:
            QuerySet[UserSurvey]: UserSurvey들에 대한 쿼리셋
        """
        survey = Survey.objects.get(**self.kwargs)
        queryset = UserSurvey.objects.filter(user=self.request.user,
                                             survey=survey).order_by('create_at')
        return queryset


class SurveyFormView(MyLoginRequiredMixin, ProcessFormView):
    """설문조사 폼을 제공하는 클래스 기반 뷰

    """

    def _create_replies(self, user_survey, post_data):
        """응답들을 생성하기 위한 메서드

        Args:
            user_survey (UserSurvey): 유저가 실시한 설문조사(새로 만들어진)
            post_data (QueryDict): 폼으로 입력받은 데이터
        """
        for key in post_data.keys():
            if not key.isdigit():  # for excepting csrf_token
                continue

            question = Question.objects.get(pk=int(key))
            survey_question = SurveyQuestion.objects.get(survey=user_survey.survey,
                                                         question=question)
            answer_content_list = post_data.getlist(key)

            answer_content = None

            # 기타 옵션 처리
            if "other" in answer_content_list:
                other_text_key = f"other_text_{key}"
                other_text = post_data.get(other_text_key)
                if other_text:
                    answer_content = other_text
                else:
                    answer_content = ''  # 기타 옵션이 선택되었지만 텍스트가 입력되지 않았을 때 빈 문자열로 설정

            # 기타 옵션이 선택되지 않았을 경우, 다른 답변 처리
            if answer_content is None or answer_content == '':
                answer_content = ','.join([item for item in answer_content_list if item != "other"])

            Reply.objects.create(user_survey=user_survey,
                                 survey_question=survey_question,
                                 content=answer_content)
            # content=','.join(post_data.getlist(key)))

    def get(self, request, *args, **kwargs):
        """폼 입력 화면

        Args:
            request (HttpRequest): HttpRequest정보

        Returns:
            HttpResponse: 렌더링 된 입력 화면 HTML
        """
        survey = Survey.objects.get(pk=kwargs['pk'])
        questions = SurveyQuestion.objects.filter(survey=survey).order_by('order')
        context = {
            'object': survey,
            'survey_questions': questions,
        }

        return render(request, 'survey/user_survey_form.html', context)

    def post(self, request, *args, **kwargs):
        """폼 제출 화면

        Args:
            request (HttpRequest): HttpRequest정보

        Returns:
            HttpResponse: 렌더링 된 완료 화면 HTML
        """
        survey = Survey.objects.get(pk=kwargs['pk'])
        user_survey = UserSurvey.objects.create(user=request.user,
                                                survey=survey)

        self._create_replies(user_survey, request.POST)
        result = ""
        total_score = None

        if "임신스트레스 10문항" in user_survey.survey_name or "조기진통위험 10문항" in user_survey.survey_name:
            replies = Reply.objects.filter(user_survey=user_survey).order_by("survey_question__order")
            scores = []
            for reply in replies:
                scores.append(Answer.objects.filter(description=reply.content).get().value)
            if "임신스트레스 10문항" in user_survey.survey_name:
                result = utils.stress_result(tuple(scores))
            else:
                result = utils.pbras_result(tuple(scores))
            
            total_score = sum(scores)
            user_survey.score = total_score
            user_survey.save()

        return render(request, 'survey/survey_complete.html', {'result': result, 'total_score': total_score})

    def put(self, request, *args, **kwargs):
        """작성된 설문 수정 화면
        
        Args:
            request (HttpRequest): HttpRequest정보
        
        Returns:
            HttpResponse: 렌더링 된 완료 화면 HTML
        """
        return render(request, 'survey/survey_complete.html')


class XlsxDownloadView(SuperuserRequiredMixin, View):

    def _create_workbook(self, user_surveys: QuerySet, user):
        wb = Workbook()
        ws = wb.active

        # create main sheet
        ws.title = '대상정보'
        ws.append(['이름', 'ID'])
        # 이름 결정: huami.fullname → fitbit.full_name → username
        if hasattr(user, 'huami') and user.huami and user.huami.fullname:
            name = user.huami.fullname
        elif hasattr(user, 'fitbit') and user.fitbit and user.fitbit.full_name:
            name = user.fitbit.full_name
        else:
            name = user.username

        ws.append([name, user.username])
        
        # 설문별 점수 라벨 간단 매핑
        def score_label(title: str):
            if "임신스트레스 10문항" in title:
                return "스트레스 점수"
            if "조기진통위험 10문항" in title:
                return "조기진통 점수"
            if "QUIPP" in title:
                return "QUIPP 점수"
            return None  # 그 외 설문은 점수 컬럼 없음
        
        # 시트명 간단 정리: "[상시] ..." 제거
        def sheet_name_from(title: str):
            m = re.search(r'\[.*?\]\s*(.*)', title)
            return m.group(1) if m else title
        
        # 3) 질문 타이틀 불러오기
        def question_titles(survey):
            qs = survey.questions.all()
            try:
                qs = qs.order_by('order')
            except FieldError:
                qs = qs.order_by('id')
            return list(qs.values_list('title', flat=True))

        for us in user_surveys:
            survey = us.survey
            sname = sheet_name_from(survey.title)
            titles = question_titles(survey)
            s_label = score_label(survey.title)

            # 시트가 없으면 헤더 만들기
            if sname not in wb.sheetnames:
                ws = wb.create_sheet(title=sname)
                header = ['작성시간', *titles]
                if s_label:
                    header.append(s_label)  # 해당 설문만 점수 컬럼 추가
                ws.append(header)
            else:
                ws = wb[sname]

            # 응답 매핑 후 행 추가
            reply_dict = {r.survey_question.question.title: r.content for r in us.replies.all()}
            row = [us.create_at, *[reply_dict.get(t, '') for t in titles]]
            if s_label:
                row.append(us.score)  # 해당 설문만 점수 값 추가
            ws.append(row)

        return wb

    def get(self, request, user_id):
        user_surveys = UserSurvey.objects.filter(user_id=user_id).order_by('survey_id', 'create_at')
        user = get_user_model().objects.get(pk=user_id)

        wb = self._create_workbook(user_surveys, user)

        with NamedTemporaryFile(suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        # 파일 이름 결정
        if hasattr(user, 'huami') and user.huami and user.huami.fullname:
            name = user.huami.fullname
        elif hasattr(user, 'fitbit') and user.fitbit and user.fitbit.full_name:
            name = user.fitbit.full_name
        else:
            name = user.username

        filename = urllib.parse.quote(f"{name} 설문결과")

        response = HttpResponse(content=stream,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

